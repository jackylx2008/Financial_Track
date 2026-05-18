from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
EXPENSE_FILL = PatternFill("solid", fgColor="FCE4D6")
INCOME_FILL = PatternFill("solid", fgColor="E2F0D9")
TRANSFER_FILL = PatternFill("solid", fgColor="D9EAF7")
REFUND_FILL = PatternFill("solid", fgColor="EADCF8")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))


def build_ledger_review_workbook(output_path: Path, entries: list[dict[str, Any]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    sorted_entries = sorted(entries, key=_sort_key)
    _write_summary(workbook, sorted_entries)

    by_month: dict[str, list[dict[str, Any]]] = defaultdict(list)
    missing_date = []
    for entry in sorted_entries:
        month = entry.get("year_month") or ""
        if month:
            by_month[month].append(entry)
        else:
            missing_date.append(entry)

    for month in sorted(by_month):
        _write_entries_sheet(workbook, month, by_month[month])
    if missing_date:
        _write_entries_sheet(workbook, "缺日期", missing_date)

    workbook.save(output_path)


def _write_summary(workbook: Workbook, entries: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("总览")
    sheet.sheet_properties.tabColor = "1F4E78"
    by_month = Counter(entry.get("year_month") or "缺日期" for entry in entries)
    by_type = Counter(entry.get("transaction_type") or "未知" for entry in entries)
    review_count = sum(1 for entry in entries if entry.get("review_required"))
    rows = [
        ["指标", "数量", "说明"],
        ["账本记录", len(entries), "ledger_entries 总记录数"],
        ["需人工复核", review_count, "review_required=true"],
        ["已匹配订单", sum(1 for entry in entries if entry.get("matched_order")), "matched_order=true"],
        ["缺日期", by_month.get("缺日期", 0), "date 为空"],
        [],
        ["收支类型", "数量", ""],
    ]
    rows.extend([[key, count, ""] for key, count in sorted(by_type.items())])
    rows.extend([[], ["月份", "数量", ""]])
    rows.extend([[key, count, ""] for key, count in sorted(by_month.items())])
    _write_matrix(sheet, rows)
    _style_sheet(sheet, table_name=None)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 48


def _write_entries_sheet(workbook: Workbook, title: str, entries: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(_safe_sheet_title(title))
    sheet.sheet_properties.tabColor = "FFC000" if title == "缺日期" else "5B9BD5"
    rows = [_entry_row(entry) for entry in sorted(entries, key=_sort_key)]
    columns = _entry_columns()
    _write_matrix(sheet, [columns] + rows)
    _style_sheet(sheet, table_name=_safe_table_name(title), rows=len(rows) + 1, cols=len(columns))


def _entry_columns() -> list[str]:
    return [
        "日期",
        "金额",
        "收支类型",
        "一级分类",
        "二级分类",
        "三级分类",
        "目标人",
        "项目",
        "标签",
        "商户/对象",
        "支付方式",
        "原始摘要",
        "订单补充明细",
        "分类理由",
        "分类置信度",
        "需复核",
        "问题提示",
        "账本ID",
        "来源流水ID",
        "人工收支类型",
        "人工一级分类",
        "人工二级分类",
        "人工三级分类",
        "人工目标人",
        "人工项目",
        "人工标签",
        "人工是否报销",
        "人工预算状态",
        "人工备注",
    ]


def _entry_row(entry: dict[str, Any]) -> list[Any]:
    merchant = entry.get("merchant") or entry.get("counterparty") or ""
    return [
        entry.get("date", ""),
        _number(entry.get("amount", "")),
        entry.get("transaction_type", ""),
        entry.get("category_lv1", ""),
        entry.get("category_lv2", ""),
        entry.get("category_lv3", ""),
        entry.get("target_person", ""),
        entry.get("project", ""),
        ", ".join(entry.get("tags", [])),
        merchant,
        entry.get("payment_method", ""),
        entry.get("raw_description", ""),
        entry.get("order_detail_summary", ""),
        entry.get("classification_reason", ""),
        entry.get("classification_confidence", ""),
        "是" if entry.get("review_required") else "否",
        ", ".join(entry.get("warnings", [])),
        entry.get("ledger_entry_id", ""),
        entry.get("source_financial_transaction_id", ""),
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]


def _write_matrix(sheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_sheet(sheet: Any, table_name: str | None, rows: int | None = None, cols: int | None = None) -> None:
    max_row = rows or sheet.max_row
    max_col = cols or sheet.max_column
    if max_row < 1 or max_col < 1:
        return
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    sheet.sheet_view.showGridLines = False

    headers = [sheet.cell(1, col).value for col in range(1, max_col + 1)]
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 28

    manual_start = headers.index("人工收支类型") + 1 if "人工收支类型" in headers else None
    type_col = headers.index("收支类型") + 1 if "收支类型" in headers else None
    review_col = headers.index("需复核") + 1 if "需复核" in headers else None
    amount_col = headers.index("金额") + 1 if "金额" in headers else None

    for row_idx in range(2, max_row + 1):
        row_fill = _row_fill(sheet.cell(row_idx, type_col).value if type_col else "")
        if review_col and sheet.cell(row_idx, review_col).value == "是":
            row_fill = REVIEW_FILL
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row_idx, col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_fill:
                cell.fill = row_fill
            if manual_start and col_idx >= manual_start:
                cell.fill = REVIEW_FILL
            if amount_col and col_idx == amount_col and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    _fit_columns(sheet, max_row, max_col)
    if table_name and max_row >= 2:
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)


def _row_fill(transaction_type: str) -> PatternFill | None:
    if transaction_type == "支出":
        return EXPENSE_FILL
    if transaction_type == "收入":
        return INCOME_FILL
    if transaction_type == "转账":
        return TRANSFER_FILL
    if transaction_type == "退款":
        return REFUND_FILL
    return None


def _fit_columns(sheet: Any, max_row: int, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        header = str(sheet.cell(1, col_idx).value or "")
        max_len = len(header)
        for row_idx in range(2, min(max_row, 200) + 1):
            value = sheet.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 80))
        if any(token in header for token in ["摘要", "明细", "理由", "问题"]):
            width = min(max(max_len + 2, 24), 58)
        elif "ID" in header or "id" in header:
            width = min(max(max_len + 2, 18), 34)
        elif header.startswith("人工"):
            width = min(max(max_len + 2, 14), 24)
        else:
            width = min(max(max_len + 2, 10), 28)
        sheet.column_dimensions[letter].width = width


def _sort_key(entry: dict[str, Any]) -> tuple[str, str, str]:
    date = entry.get("date") or "9999-99-99"
    return (date, str(entry.get("transaction_type", "")), str(entry.get("ledger_entry_id", "")))


def _number(value: Any) -> Any:
    if value in {None, ""}:
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _safe_sheet_title(value: str) -> str:
    title = str(value or "Sheet").replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "")
    title = title.replace("[", "").replace("]", "").replace(":", "-")
    return title[:31] or "Sheet"


def _safe_table_name(value: str) -> str:
    base = "LedgerReview"
    text = str(value or "")
    if _is_month(text):
        return f"{base}{text.replace('-', '')}"
    if text == "缺日期":
        return f"{base}MissingDate"
    return base


def _is_month(value: str) -> bool:
    try:
        datetime.strptime(value, "%Y-%m")
        return True
    except ValueError:
        return False
