from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
LINKED_FILL = PatternFill("solid", fgColor="E2F0D9")
CANDIDATE_FILL = PatternFill("solid", fgColor="FFF2CC")
UNMATCHED_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))


def build_review_workbook(
    *,
    output_path: Path,
    facts: list[dict[str, Any]],
    links: list[dict[str, Any]],
    orders: list[dict[str, Any]],
    bank_transactions: list[dict[str, Any]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    order_facts = [item for item in facts if item.get("fact_type") == "order_purchase"]
    bank_payment_facts = [item for item in facts if item.get("fact_type") == "bank_payment"]
    linked = [item for item in links if item.get("match_strength") == "linked"]
    candidates = [item for item in links if item.get("match_strength") == "candidate"]
    unmatched_orders = [item for item in order_facts if item.get("link_status") == "unmatched"]

    _write_summary(workbook, facts, links, order_facts)
    _write_rows(workbook, "强关联", _link_columns(), [_link_row(item) for item in linked], tab_color="70AD47")
    _write_rows(workbook, "候选关联", _link_columns(), [_link_row(item) for item in candidates], tab_color="FFC000")
    _write_rows(
        workbook,
        "未匹配订单",
        _fact_columns(include_link=False),
        [_fact_row(item, include_link=False) for item in unmatched_orders],
        tab_color="ED7D31",
    )
    _write_rows(workbook, "订单明细", _order_columns(), [_order_row(item) for item in orders], tab_color="5B9BD5")
    _write_rows(
        workbook,
        "银行付款",
        _bank_columns(),
        [_bank_row(item) for item in bank_transactions if item.get("direction") == "outflow"],
        tab_color="4472C4",
    )
    _write_rows(workbook, "财务事实", _fact_columns(include_link=True), [_fact_row(item, include_link=True) for item in facts], tab_color="A5A5A5")

    workbook.save(output_path)


def _write_summary(workbook: Workbook, facts: list[dict[str, Any]], links: list[dict[str, Any]], order_facts: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("总览")
    sheet.sheet_properties.tabColor = "1F4E78"
    by_fact = Counter(item.get("fact_type", "unknown") for item in facts)
    by_business = Counter(item.get("business_type", "unknown") for item in facts)
    by_link = Counter(item.get("match_strength", "unknown") for item in links)
    by_order_status = Counter(item.get("link_status", "unknown") for item in order_facts)
    rows = [
        ["指标", "数量", "说明"],
        ["银行 normalized", by_fact.get("bank_payment", 0) + by_fact.get("bank_money_movement", 0), "银行流水归一化记录数"],
        ["订单 normalized", by_fact.get("order_purchase", 0), "平台订单归一化记录数"],
        ["财务事实表", len(facts), "financial_transactions 总记录数"],
        ["订单-付款关联", len(links), "financial_transaction_links 总记录数"],
        ["强关联", by_link.get("linked", 0), "match_strength=linked"],
        ["候选关联", by_link.get("candidate", 0), "match_strength=candidate"],
        ["订单 linked", by_order_status.get("linked", 0), "订单事实已强关联付款"],
        ["订单 candidate", by_order_status.get("candidate", 0), "订单事实存在候选付款"],
        ["订单 unmatched", by_order_status.get("unmatched", 0), "订单事实未匹配付款"],
        [],
        ["业务类型", "数量", ""],
    ]
    rows.extend([[key, count, ""] for key, count in sorted(by_business.items())])
    _write_matrix(sheet, rows)
    _style_sheet(sheet, table_name=None)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 48


def _write_rows(workbook: Workbook, title: str, columns: list[str], rows: list[list[Any]], tab_color: str) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_properties.tabColor = tab_color
    _write_matrix(sheet, [columns] + rows)
    _style_sheet(sheet, table_name=_safe_table_name(title), rows=len(rows) + 1, cols=len(columns))


def _write_matrix(sheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_sheet(sheet: Any, table_name: str | None, rows: int | None = None, cols: int | None = None) -> None:
    max_row = rows or sheet.max_row
    max_col = cols or sheet.max_column
    if max_row < 1 or max_col < 1:
        return
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    sheet.sheet_view.showGridLines = False

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 28

    for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00" if isinstance(cell.value, float) else "#,##0"

    _apply_status_fills(sheet, max_row, max_col)
    _fit_columns(sheet, max_row, max_col)
    if table_name and max_row >= 2:
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)


def _fit_columns(sheet: Any, max_row: int, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        header = str(sheet.cell(1, col_idx).value or "")
        max_len = len(header)
        for row_idx in range(2, min(max_row, 200) + 1):
            value = sheet.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 80))
        if any(token in header for token in ["摘要", "标题", "证据", "说明", "来源"]):
            width = min(max(max_len + 2, 24), 58)
        elif any(token in header for token in ["ID", "id"]):
            width = min(max(max_len + 2, 18), 34)
        else:
            width = min(max(max_len + 2, 10), 28)
        sheet.column_dimensions[letter].width = width


def _apply_status_fills(sheet: Any, max_row: int, max_col: int) -> None:
    headers = [sheet.cell(1, col).value for col in range(1, max_col + 1)]
    for key in ["关联状态", "强度"]:
        if key not in headers:
            continue
        col_idx = headers.index(key) + 1
        for row_idx in range(2, max_row + 1):
            value = sheet.cell(row_idx, col_idx).value
            fill = None
            if value == "linked":
                fill = LINKED_FILL
            elif value == "candidate":
                fill = CANDIDATE_FILL
            elif value in {"unmatched", "unlinked"}:
                fill = UNMATCHED_FILL
            if fill:
                for col in range(1, max_col + 1):
                    sheet.cell(row_idx, col).fill = fill


def _link_columns() -> list[str]:
    return ["强度", "分数", "金额", "平台", "订单时间", "付款时间", "订单摘要", "付款摘要", "证据", "订单ID", "银行流水ID", "关联ID"]


def _link_row(item: dict[str, Any]) -> list[Any]:
    return [
        item.get("match_strength", ""),
        item.get("score", ""),
        _number(item.get("amount", "")),
        item.get("platform", ""),
        item.get("order_time", ""),
        item.get("payment_time", ""),
        item.get("order_summary", ""),
        item.get("payment_summary", ""),
        ", ".join(item.get("evidence", [])),
        item.get("order_record_id", ""),
        item.get("bank_transaction_id", ""),
        item.get("link_id", ""),
    ]


def _fact_columns(*, include_link: bool) -> list[str]:
    base = ["类型", "业务类型", "时间", "金额", "方向", "平台", "支付渠道", "商户", "对方", "标题", "摘要", "置信度", "警告", "财务事实ID"]
    if include_link:
        base[1:1] = ["关联状态", "最佳分数", "候选数"]
    return base


def _fact_row(item: dict[str, Any], *, include_link: bool) -> list[Any]:
    row = [
        item.get("fact_type", ""),
        item.get("business_type", ""),
        item.get("occurrence_time", ""),
        _number(item.get("amount", "")),
        item.get("direction", ""),
        item.get("platform", ""),
        item.get("payment_channel", ""),
        item.get("merchant", ""),
        item.get("counterparty", ""),
        item.get("title", ""),
        item.get("summary", ""),
        item.get("confidence", ""),
        ", ".join(item.get("warnings", [])),
        item.get("financial_transaction_id", ""),
    ]
    if include_link:
        row[1:1] = [item.get("link_status", ""), item.get("best_link_score", ""), item.get("link_candidate_count", "")]
    return row


def _order_columns() -> list[str]:
    return ["平台", "订单时间", "金额", "商户", "标题", "规格", "数量", "状态", "关联提示", "置信度", "警告", "来源图片", "订单ID", "订单记录ID"]


def _order_row(item: dict[str, Any]) -> list[Any]:
    return [
        item.get("platform", ""),
        item.get("order_time", ""),
        _number(item.get("paid_amount", "")),
        item.get("merchant", ""),
        item.get("title", ""),
        item.get("spec", ""),
        item.get("quantity", ""),
        item.get("status", ""),
        "需复核" if item.get("warnings") else "",
        item.get("confidence", ""),
        ", ".join(item.get("warnings", [])),
        item.get("source_image", ""),
        item.get("order_id", ""),
        item.get("order_record_id", ""),
    ]


def _bank_columns() -> list[str]:
    return ["时间", "金额", "银行", "账户", "商户", "对方", "摘要", "渠道", "置信度", "警告", "银行流水ID"]


def _bank_row(item: dict[str, Any]) -> list[Any]:
    return [
        item.get("transaction_time") or item.get("posting_date", ""),
        _number(item.get("amount", "")),
        item.get("bank_name") or item.get("bank_key", ""),
        item.get("account_key", ""),
        item.get("merchant", ""),
        item.get("counterparty", ""),
        item.get("summary", ""),
        item.get("channel", ""),
        item.get("confidence", ""),
        ", ".join(item.get("warnings", [])),
        item.get("transaction_id", ""),
    ]


def _number(value: Any) -> Any:
    if value in {None, ""}:
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _safe_table_name(value: str) -> str:
    mapping = {
        "强关联": "LinkedTable",
        "候选关联": "CandidateTable",
        "未匹配订单": "UnmatchedOrdersTable",
        "订单明细": "OrdersTable",
        "银行付款": "BankPaymentsTable",
        "财务事实": "FinancialFactsTable",
    }
    return mapping.get(value, "ReviewTable")
